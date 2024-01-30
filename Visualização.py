import sys
import pandas as pd
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from PyQt5.QtGui import QColor, QFont
from PyQt5.QtWidgets import QApplication, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, QLabel, QHeaderView

class FixedSizeTableWidget(QTableWidget):
    def __init__(self, rows, columns, cell_width, cell_height):
        super().__init__(rows, columns)
        self.cell_width = cell_width
        self.cell_height = cell_height
        self.initUI()
    
    def initUI(self):
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        
        for i in range(self.columnCount()):
            self.setColumnWidth(i, self.cell_width)
            
        for i in range(self.rowCount()):
            self.setRowHeight(i, self.cell_height)
            
class TableExample(QWidget):
    def __init__(self):
        super().__init__()

        self.initUI()

    def determinar_cor(self, valor, coluna, opacidade=150):
        if coluna >= 210:  # Se a coluna estiver na faixa de colunas extras
            return QColor(255, 255, 255, opacidade)  # Branco
        elif valor == "ABQ3 - Padrão 2,25":
            return QColor(0, 176, 240, opacidade)  # Azul claro
        elif valor == "ABQ3 - Padrão 2,26":
            return QColor(255, 255, 0, opacidade)  # Amarelo
        elif valor == "ABQ3 - Padrão 2,40":
            return QColor(255, 0, 0, opacidade)  # Vermelho
        elif valor == "ABQ3 - Padrão Estreito":
            return QColor(112, 48, 160, opacidade)  # Roxo
        elif valor == "ABQ3 - Padrão 2,00":
            return QColor(0, 0, 0, opacidade)  # Preto
        elif valor == "ABQ3 - Não Padrão":
            return QColor(0, 176, 80, opacidade)  # Verde escuro
        elif valor == "ABQ3 - BQD":
            return QColor(255, 192, 0, opacidade)  # Laranja
        elif valor == "ABQ3 - BQD UPV":
            return QColor(0, 255, 0, opacidade)  # Verde claro
        elif valor == "Descanso/Retrabalho":
            return QColor(255, 102, 255, opacidade)  # Rosa
        elif coluna % 5 == 4:
            return QColor(211, 211, 211, opacidade)
        else:
            return QColor(255, 255, 255, opacidade)  # Branco
        

    def initUI(self):
        self.setWindowTitle('Visualização de estoque')
        self.setGeometry(200, 200, 900, 900)

        excel_file_path = 'LX03.xlsx'
        df = pd.read_excel(excel_file_path)

        positions = df['Posição'].tolist()
        lotes = df['Lote'].tolist()
        classifications = df['Classificação'].tolist()
        situacoes_lote = df['Situação do lote'].tolist()

        self.layout = QVBoxLayout()

        title_label = QLabel('Visualização ABQ3')
        title_font = QFont("Courier New", 14, QFont.Bold)
        title_label.setFont(title_font)
        self.layout.addWidget(title_label)
        title_label.setAlignment(Qt.AlignCenter)

        # Primeira tabela
        self.tableWidget1 = FixedSizeTableWidget(9, 210, 100, 40)
        
        level_mapping = {1: 4, 2: 3, 3: 2, 4: 1}

        self.tableWidget1.horizontalHeader().setStyleSheet("""
            QHeaderView::section {
                background-color: rgb(211, 211, 211);
            }
        """)

        # Configurações e preenchimento da primeira tabela...
        header_labels = [f'{i // 5 + 1}.{level_mapping[i % 5 + 1]}' if i % 5 != 4 else '' for i in range(210)]
        self.tableWidget1.setHorizontalHeaderLabels(header_labels)

        vertical_header_labels = ['A', 'B', 'C', 'D', 'E', 'F', ' ', 'Rack', 'Vazios']
        self.tableWidget1.setVerticalHeaderLabels(vertical_header_labels)

        for position, lote, classification, situacao_lote in zip(positions, lotes, classifications, situacoes_lote):
            col = (int(position[:2]) - 1) * 5
            row = ord(position[2]) - ord('A')
            level = level_mapping[int(position[3])] - 1
            
            print(f"Posição: {position}, Lote: {lote}, Coord: ({row}, {col + level})")

            item = QTableWidgetItem(f"{lote}\n{situacao_lote}")
            item.setBackground(QColor(self.determinar_cor(classification, col + level)))
            self.tableWidget1.setItem(row, col + level, item)

        emptyCellColor = QColor(255, 255, 255)

        emptyCellCounts = []  # Adicione esta linha para criar uma lista vazia
        # Adicione o número de células vazias em cada rack à lista emptyCellCounts
        for rack in range(42):  # Para cada rack
            emptyCellCount = 0
            for row in range(6):  # Para cada linha no rack
                for level in range(4):  # Para cada coluna no rack
                    column = rack * 5 + level
                    if column < self.tableWidget1.columnCount():
                        item = self.tableWidget1.item(row, column)
                        if item is None or item.background().color() == emptyCellColor:
                            emptyCellCount += 1

            #print(f"Number of empty cells in rack {rack + 1}: {emptyCellCount}")
            emptyCellCounts.append(emptyCellCount)  # Adicione o resultado à lista emptyCellCounts
        
        # Adicione o número total de células vazias à tabela
        for rack in range(42):
            emptyCellItem = QTableWidgetItem(f"{emptyCellCounts[rack]}")
            emptyCellItem.setTextAlignment(Qt.AlignCenter)
            self.tableWidget1.setItem(self.tableWidget1.rowCount() - 1, rack * 5, emptyCellItem)
                       
        df_empty_cells = pd.DataFrame({
            'Rack': list(range(1, 43)),
            'Empty_Cells_Count': emptyCellCounts
            })
        
        excel_output_path = 'empty_cells_count.xlsx'
        df_empty_cells.to_excel(excel_output_path, index=False)
        
        try:
            with pd.ExcelWriter(excel_output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_empty_cells.to_excel(writer, sheet_name='Sheet1', index=False)
            
            #print(f"DataFrame salvo e coluna 'Empty_Cells_Count' atualizada em {excel_output_path}")
        
        except FileNotFoundError:
            print(f"Arquivo {excel_output_path} não encontrado.")
            
        except Exception as e:
            print(f"Erro ao atualizar coluna 'Empty_Cells_Count': {e}")

       # Carregar cores do arquivo xlsx para a tabela
        self.load_colors_from_xlsx('cores_quadriplicadas.xlsx')

        for i in range(4, 210, 5):
            for j in range(9):
                item = QTableWidgetItem('')
                item.setBackground(QColor(211, 211, 211))
                self.tableWidget1.setItem(j, i, item)

        self.tableWidget1.resizeColumnsToContents()

        max_width = max(self.tableWidget1.columnWidth(i) for i in range(self.tableWidget1.columnCount()))

        for i in range(self.tableWidget1.columnCount()):
            self.tableWidget1.setColumnWidth(i, max_width)


        # Configuração do layout
        self.layout.addWidget(self.tableWidget1, 2)
        
        legend_label = QLabel('Legenda:')
        self.layout.addWidget(legend_label)

        
        legend = QVBoxLayout()
        legend_items = [
            ('ABQ3 - Padrão 2,25', 'Azul claro', '#00B0F0'),
            ('ABQ3 - Padrão 2,26', 'Amarelo', '#FFFF00'),
            ('ABQ3 - Padrão 2,40', 'Vermelho', '#FF0000'),
            ('ABQ3 - Padrão Estreito', 'Roxo', '#7030A0'),
            ('ABQ3 - Padrão 2,00', 'Preto', '#000000'),
            ('ABQ3 - Não Padrão', 'Verde escuro', '#00B050'),
            ('ABQ3 - BQD', 'Laranja', '#FFC000'),
            ('ABQ3 - BQD UPV', 'Verde claro', '#00FF00'),
            ('Descanso/Retrabalho', 'Rosa', '#FF66FF'),
        ]
        
        for item in legend_items:
            legend_item_label = QLabel(f"{item[0]}: {item[1]}")
            legend_item_label.setStyleSheet(f"background-color: {item[2]};")
            legend.addWidget(legend_item_label)
        
        self.layout.addLayout(legend)
        
        self.setLayout(self.layout)
        
        print("Terminou de carregar dados do Excel...")


        self.show()
        
    

    def load_colors_from_xlsx(self, file_path):
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            row = 7

            for col, cell in enumerate(sheet.iter_rows(values_only=True)):
                for color_str in cell:
                    # Certifique-se de que a string de cor comece com '#'
                    if color_str.startswith('#'):
                        color_str = color_str[1:]  # Remova o caractere '#' inicial

                    # Garanta que a string de cor tem 6 caracteres (formato hexadecimal completo)
                    color_str = color_str.ljust(6, '0')[:6]

                    # Converta a string de cor para uma tupla de inteiros
                    color_rgb = tuple(int(color_str[i:i + 2], 16) for i in (0, 2, 4))
                    item = QTableWidgetItem('')
                    item.setBackground(QColor(*color_rgb))
                    self.tableWidget1.setItem(row, col, item)

        except FileNotFoundError:
            print(f"Arquivo {file_path} não encontrado.")
        except Exception as e:
            print(f"Erro ao carregar cores do arquivo xlsx: {e}")

    def load_colors_from_xlsx_configuracao(self, file_path, table_widget, row_index):
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            classificacoes = [
                "ABQ3 - Padrão 2,25",
                "ABQ3 - Padrão 2,26",
                "ABQ3 - Padrão 2,40",
                "ABQ3 - Padrão Estreito",
                "ABQ3 - Padrão 2,00",
                "ABQ3 - Não Padrão",
                "ABQ3 - BQD",
                "ABQ3 - BQD UPV",
                "Descanso/Retrabalho"
            ]

            for col, row in enumerate(sheet.iter_rows(min_col=1, max_col=42, min_row=row_index, max_row=row_index)):
                for cell, classificacao in zip(row, classificacoes):
                    color_str = cell.value
                    if color_str is not None:
                        # Certifique-se de que a string de cor comece com '#'
                        if color_str.startswith('#'):
                            color_str = color_str[1:]  # Remova o caractere '#' inicial

                        # Garanta que a string de cor tem 6 caracteres (formato hexadecimal completo)
                        color_str = color_str.ljust(6, '0')[:6]

                        # Converta a string de cor para uma tupla de inteiros
                        color_rgb = tuple(int(color_str[i:i + 2], 16) for i in (0, 2, 4))
                        item = QTableWidgetItem('')
                        item.setBackground(QColor(*color_rgb))
                        table_widget.setItem(row_index, col * 2, item)

                        # Adicione a classificação ao lado da cor
                        classificacao_item = QTableWidgetItem(classificacao)
                        table_widget.setItem(row_index, col * 2 + 1, classificacao_item)

        except FileNotFoundError:
            print(f"Arquivo {file_path} não encontrado.")
        except Exception as e:
            print(f"Erro ao carregar cores do arquivo xlsx: {e}")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = TableExample()
    sys.exit(app.exec_())
